# coding=utf-8

from openpyxl import workbook, load_workbook
from collections import Counter
import wget
import requests
import re
import time
import os

manual_case = [
               'TC_USB_AaaG_033',
               'TC_USB_AaaG_035',
               'TC_USB_AaaG_013',
               'TC_USB_AaaG_015',
               'TC_USB_Mediator_AaaG_007',
               'TC_USB_Mediator_AaaG_022',
               'TC_Virtio_AaaG_028',
               'TC_Virtio_AaaG_066',
               'TC_USB_SOS_015'
               ]


def rdexcel():
    wb = load_workbook(filename='test.xlsx')
    sheets = wb.sheetnames
    print(sheets)
    sheet_first = sheets[0]
    ws = wb['ww44.5']
    rows = ws.rows
    columns = ws.columns
    print(ws.max_row, ws.max_column)
    num_case = list()
    name_column = list()
    for column in columns:
        # print(column[0].value)
        if column[0].value == "name":
            name_column = list(column)
        if column[0].value == "status":
            # for i, status in enumerate(column):
            #     if status.value == "FAIL":
            #         num_case.append(i)
            num_case = [i for i, status in enumerate(column) if status.value == "FAIL"]
            print(num_case)
        if num_case and name_column:
            print("case_name".center(50, "#"))
            case_name = [name.value for i, name in enumerate(name_column) if i in num_case]
            for i in range(len(manual_case)):
                if manual_case[i] in case_name[:]:
                    case_name.remove(manual_case[i])
            print(list(case_name))
            print(len(case_name))
            break
            # print(column)
    # print(ws['A1'].value)
    # wb.create_sheet('test', 0)
    # wb.remove(wb[sheets[0]])
    print(sheets)
    # wb.save('test.xlsx')
    wb.close()

def downxlsxfile():
    local_addr = 'root@10.239.147.17'
    file_addr = '/var/www/html/ftp/mrb/'
    date_dir = os.popen('ssh %s %s'%(local_addr, 'ls -Ft /var/www/html/ftp/mrb | grep /$ | head -n 1')).read().strip()
    cmd = 'ls -t %s | grep xlsx$'% (file_addr+date_dir)
    xlsx_name = os.popen('ssh %s %s'%(local_addr, cmd)).read().strip()
    os.system('scp %s:%s%s%s .' % (local_addr, file_addr, date_dir, xlsx_name))

def wtexcel():
    wb = load_workbook('test.xlsx')


downxlsxfile()